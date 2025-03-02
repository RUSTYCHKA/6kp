import os
import time
import pandas as pd
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager

from datetime import datetime




def setup_driver():
    options = Options()
   # options.add_argument("--headless")  
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")   
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    return driver

def find_element_by_tag_or_param(driver, tag, param, xpath = None, inclass=None):
    
    try:
        if 'inclass=' in param:
            inclass = param.split('=')[1]
            return driver.find_element(By.XPATH, f"//*[contains(text(), '{inclass}')]" )
        elif '//' == param[:2]:
            return driver.find_element(By.XPATH, param)
        elif pd.notna(tag) and pd.notna(param):
            return driver.find_element(By.XPATH, f"//{tag}[@{param.split('=')[0]}='{param.split('=')[1]}']")
        elif pd.notna(tag):
            return driver.find_element(By.TAG_NAME, tag)
        elif pd.notna(param):
            return driver.find_element(By.XPATH, f"//*[@{param.split('=')[0]}='{param.split('=')[1]}']")

    except:
        return None





async def submit_form(driver, row, user_name, user_phone, user_email, message, path):
    driver.get(row['URL'])
    time.sleep(3)  
    
    
    try:
        search = find_element_by_tag_or_param(driver, row['Кнопка оставить заявку, если по ссылке сразу нельзя оставить заявку(тег)'], row['Кнопка оставить заявку, если по ссылке сразу нельзя оставить заявку(параметр)'])
        
        if search:
            search.click()
        
        name_input = find_element_by_tag_or_param(driver, row['Имя(тег)'], row['Имя(параметр)'])
        phone_input = find_element_by_tag_or_param(driver, row['Телефон(тег)'], row['Телефон(параметр)'])
        email_input = find_element_by_tag_or_param(driver, row['Почта(тег)'], row['Почта(параметр)'])
        
        try:
            if name_input:
                name_input.send_keys(user_name)
            if phone_input:
                phone_input.send_keys(user_phone)
            if email_input:
                email_input.send_keys(user_email)
        except ElementNotInteractableException:
            if name_input:
                name_input.click()
                name_input.send_keys(user_name)    
            if phone_input:
                phone_input.click()
                phone_input.send_keys(user_phone)
            if email_input:
                email_input.click()
                email_input.send_keys(user_email)
        
        
        submit_button = find_element_by_tag_or_param(driver, row['Кнопка отправки заявки, оставить пустой, если достаточно после заполнения нажать Enter(тег)'], row['Кнопка отправки заявки, оставить пустой, если достаточно после заполнения нажать Enter(параметр)'])
        
        if submit_button:
            submit_button.click()
        else:
            if name_input:
                name_input.send_keys(Keys.ENTER)
            if phone_input:
                phone_input.send_keys(Keys.ENTER)
            if email_input:
                email_input.send_keys(Keys.ENTER)
        
        time.sleep(2)
        add_data_to_excel(path, row['URL'], 'Успешно')
        #await message.answer(f"Успешно отправлена заявка на {row['URL']}")
        return True
    except Exception as e:
        #await message.answer(f"Ошибка на {row['URL']}: {e}")
        add_data_to_excel(path, row['URL'], f"Ошибка: {e}")
        return False




def create_excel_file(filename):
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    
    sheet['A1'] = 'Ccылка на сайт'
    sheet['B1'] = 'Cтатус'

   
    workbook.save(filename)
    print(f"Файл '{filename}' создан.")




def add_data_to_excel(filename, url, status):
    
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    
    next_row = sheet.max_row + 1

    
    sheet[f'A{next_row}'] = url
    sheet[f'B{next_row}'] = status

    
    workbook.save(filename)
    print(f"Данные добавлены: {url}, {status}")


async def submit_forms(user_name, user_phone, user_email, message, category):

    
    driver = setup_driver()
    success_count = 0
    if not os.path.exists(f'tables/{message.from_user.id}'):
        os.mkdir(f'tables/{message.from_user.id}')
    path = f'tables/{message.from_user.id}/{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    create_excel_file(path)
    if category is not None:
        data = pd.read_excel("sites_data.xlsx", sheet_name=category )
    else:
        data = pd.read_excel("sites_data.xlsx")
        
    for _, row in data.iterrows():
        res =  await submit_form(driver, row, user_name, user_phone, user_email, message, path)
        if res:
            success_count += 1
    
    driver.quit()

